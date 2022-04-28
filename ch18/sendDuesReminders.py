#! python3
# sendDuesReminders.py - スプレッドシートの支払い状況に基づきメールを送信

import openpyxl, smtplib, sys

# スプレッドシートを開き最近の支払い状況を取得
wb = openpyxl.load_workbook('duesRecords.xlsx')  # ❶
sheet = wb['Sheet1'] # ❷

last_col = sheet.max_column  # ❸
latest_month = sheet.cell(row=1, column=last_col).value  # ❹

# 会員の支払い状況を調べる
unpaid_members = {}
for r in range(2, sheet.max_row + 1):        # ❶
    payment = sheet.cell(row=r, column=last_col).value # ❷
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value       # ❸
        email = sheet.cell(row=r, column=2).value      # ❹
        unpaid_members[name] = email  # ❺

# メールアカウントにログインする
smtp_obj = smtplib.SMTP('smtp.example.com', 587)
smtp_obj.ehlo()
smtp_obj.starttls()
# SSL接続の場合
# smtp_obj = smtplib.SMTP_SSL('smtp.example.com', 465)
# smtp_obj.ehlo()
smtp_obj.login('my_email_address@example.com', sys.argv[1])

# リマインダーメールを送信する
for name, email in unpaid_members.items():
    body = f"""Subject: {latest_month} dues unpaid.
Dear {name},
Records show that you have not paid dues for {latest_month}.
Please make this payment as soon as possible. Thank you!
"""  # ❶
    print(f'メール送信中 {email}...') # ❷
    sendmail_status = smtp_obj.sendmail('my_email_address@example.com', email, body) # ❸

    if sendmail_status != {}: # ❹
        print(f'{email}へメール送信中に問題が起こりました: {sendmail_status}')

smtp_obj.quit()

