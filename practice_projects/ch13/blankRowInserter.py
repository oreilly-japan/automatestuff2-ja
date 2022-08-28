# 演習プロジェクト 13.14.2  空行を挿入する

import openpyxl
import sys

if len(sys.argv) < 4:
    sys.exit("""
Usage: python blankRowInserter.py N M src.xlsx
  N = start row number (1..)
  M = the number of blank rows
""")

n = int(sys.argv[1])
m = int(sys.argv[2])
src = sys.argv[3]

wb = openpyxl.load_workbook(src)
sheet = wb.active

new_wb = openpyxl.Workbook()
new_sheet = new_wb.active

for row in range(1, sheet.max_row + 1):
    if row < n:
        new_row = row
    else:
        new_row = row + m  # m行分の空行を入れる

    for col in range(1, sheet.max_column + 1):
        old_cell = sheet.cell(column=col, row=row)
        new_cell = new_sheet.cell(column=col, row=new_row)
        new_cell.value = old_cell.value

new_wb.save(src + '.ins.xlsx')
