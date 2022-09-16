# 演習プロジェクト 13.14.1  N×Nの掛け算の表を作成する

import openpyxl
from openpyxl.styles import Font
import sys

if len(sys.argv) < 2:
    sys.exit('Usage: python multiplicationTable.py N')

n = int(sys.argv[1])

wb = openpyxl.Workbook()
sheet = wb.active
normal_font = Font(size=12)          # 通常フォント
bold_font = Font(size=12, bold=True) # 太字フォント

# 行見出し
for i in range(1, n + 1):
    cell = sheet.cell(row=i + 1, column=1)
    cell.value = i
    cell.font = bold_font

# 列見出し
for j in range(1, n + 1):
    cell = sheet.cell(row=1, column=j + 1)
    cell.value = j
    cell.font = bold_font

# 九九
for i in range(1, n + 1):
    for j in range(1, n + 1):
        cell = sheet.cell(row=i + 1, column=j + 1)
        cell.value = i * j
        cell.font = normal_font

wb.save('multiplicationtable.xlsx')
