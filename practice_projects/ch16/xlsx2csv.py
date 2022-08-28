# 演習プロジェクト 16.8.1 ExcelからCSVへ変換

import openpyxl
import sys
import os
import csv

for excel_file in os.listdir('.'):
    # xlsxファイルでなければスキップし、Workbookオブジェクトを読み込む
    if not excel_file.lower().endswith('.xlsx'):
        continue
    wb = openpyxl.load_workbook(excel_file)

    for sheet_name in wb.sheetnames:
        # ワークブックのシートをループする
        sheet = wb[sheet_name]

        # Excelファイル名とシート名からCSVファイル名を作る
        csv_filename = excel_file[:-5] + '_' + sheet_name + '.csv'
        # このCSVファイル用にcsv.writerオブジェクトを生成する
        csv_file = open(csv_filename, 'w', encoding='utf-8', newline='')
        csv_writer = csv.writer(csv_file)

        for row_num in range(1, sheet.max_row + 1):
            row_data = [] # セルをこのリストに追加する
            # 行のセルをループする
            for col_num in range(1, sheet.max_column + 1):
                 # セルをrow_dataに追加する
                 row_data.append(sheet.cell(column=col_num, row=row_num).value)
            # row_dataをCSVファイルに書き出す
            csv_writer.writerow(row_data)

        csv_file.close()

